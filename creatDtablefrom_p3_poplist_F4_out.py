
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, colors
import pandas as pd
import sys

def read_D(name):
    Dset = {}
    with open(name) as Df:
        rawline = Df.readline()
        while rawline:
            eachline = rawline.strip().split()
            p1 = eachline[1]
            p2 = eachline[2]
            p3 = eachline[3]
            p4 = eachline[4]
            d  = eachline[5]
            Z  = eachline[6]
            Dset[(p1,p2,p3,p4)]=(d,Z)   # core data
            rawline = Df.readline()
    return Dset


def read_poplist(name):
    Pf = open(name,'r')
    poplist = []
    for rawline in Pf:
        eachline = rawline.strip().split()
        poplist.append(eachline[0])
    return poplist
    Pf.close()

#pd = '/home/rjk19/data/'
P3 = sys.argv[1]
pop = read_poplist(sys.argv[2])
pop.remove(P3)
pop = list(set(pop))
print('poplist has been read')

# write the D value and color in excel

wb = Workbook()
ws = wb.active
Dset = read_D(sys.argv[3])
for indexP2,P2 in enumerate(pop):
    for indexP4,P4 in enumerate(pop):
        if P3 != P4 and P3 != P2 and P4 != P2 :
            panel = ('Mbuti',P2,P3,P4)
            Zi = Dset[panel][1]    # get the Z in the Dset
            ws.cell(row=indexP2+2, column=indexP4+2).value=float(Zi)
            if float(Zi) > 3:
                mycol = 'FF9999'
            elif float(Zi) < -3:
                mycol = '83CFF6'
            elif float(Zi) > 2.5 and float(Zi) < 3:
                mycol = 'FFCCCC'
            elif float(Zi) < -2.5 and float(Zi) > -3:
                mycol = 'CCECFF'
            else:
                mycol = 'FFFFFF'
            fill_col = PatternFill("solid", fgColor = mycol)
            ws.cell(row=indexP2+2, column=indexP4+2).value=float(Zi)  
            ws.cell(row=indexP2+2, column=indexP4+2).fill=fill_col
        else:
            ws.cell(row=indexP2+2, column=indexP4+2).value= None

# write the pop name in excel
for indexP,P in enumerate(pop):
    ws.cell(row=1, column=indexP2+2).value=P2
    ws.cell(row=indexP2+2, column=1).value=P2

wb.save('/home/rjk/result/D_%s.xlsx' % sys.argv[4])
print('table finished')
